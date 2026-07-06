"""Correcteur Excel generique, entierement pilote par un fichier de config JSON.

Aucune cellule, valeur de reference ou bareme n'est code en dur ici : tout
provient du dictionnaire `config` (charge depuis data/epreuves/*.json).
"""

import csv
import re
import zipfile
from functools import cmp_to_key
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .fonctions_multilingues import synonymes
from integrite import verifier_watermark

STATUT_REUSSI = "reussi"
STATUT_ECHOUE = "echoue"


# Doit correspondre a CELLULE_NOM / CELLULE_PRENOM dans generateur/gen_excel.py
FEUILLE_GARDE = "0 - Consignes"
CELLULE_NOM = "B3"
CELLULE_PRENOM = "B4"


def extraire_identite(nom_fichier):
    """Extrait nom/prenom/module/session depuis NOM_Prenom_Module_Session.xlsx (repli si absent de la feuille)."""
    parts = Path(nom_fichier).stem.split("_")
    if len(parts) < 4:
        return {"nom": Path(nom_fichier).stem, "prenom": "", "module": "", "session": ""}
    return {"nom": parts[0], "prenom": parts[1], "module": parts[2], "session": "_".join(parts[3:])}


def _identite_depuis_feuille(wb_valeurs):
    """Lit le Nom/Prenom encode par l'etudiant sur la feuille de garde, si presente."""
    if FEUILLE_GARDE not in wb_valeurs.sheetnames:
        return None
    ws = wb_valeurs[FEUILLE_GARDE]
    nom = ws[CELLULE_NOM].value
    prenom = ws[CELLULE_PRENOM].value
    if not nom and not prenom:
        return None
    return {"nom": str(nom or "").strip(), "prenom": str(prenom or "").strip()}


def _formule_str(cell):
    valeur = cell.value if cell is not None else None
    if isinstance(valeur, str) and valeur.startswith("="):
        return valeur[1:].upper()
    return None


def _contient_fonction(formule, fonctions_attendues):
    if formule is None:
        return False
    for fonction in fonctions_attendues:
        for syn in synonymes(fonction):
            if re.search(rf"\b{re.escape(syn)}\s*\(", formule):
                return True
    return False


def _nom_fonction_detectee(formule):
    """Extrait le nom de la premiere fonction Excel utilisee dans une formule, pour le feedback."""
    if formule is None:
        return None
    m = re.search(r"([A-Z][A-Z0-9._]*)\s*\(", formule)
    return m.group(1) if m else None


def _valeur_proche(valeur_calculee, ref, tolerance):
    if ref is None:
        return True
    if isinstance(ref, str):
        return isinstance(valeur_calculee, str) and valeur_calculee.strip().lower() == ref.strip().lower()
    try:
        return valeur_calculee is not None and abs(float(valeur_calculee) - float(ref)) <= tolerance
    except (TypeError, ValueError):
        return False


def _corriger_formule_fonction(ws_formules, ws_valeurs, critere):
    cellules = critere["cellules"]
    formule_obligatoire = critere.get("formule_obligatoire", True)
    fonctions_attendues = critere.get("fonctions_attendues", [])
    valeur_reference = critere.get("valeur_reference")
    valeurs_reference = critere.get("valeurs_reference", {})
    tolerance = critere.get("tolerance", 0.01)

    points_max = critere.get("points_par_cellule", 0) * len(cellules) if critere.get("points_par_cellule") else critere.get("points", len(cellules))
    points_par_cellule = points_max / len(cellules)

    points_obtenus = 0.0
    details = []
    for cellule in cellules:
        formule = _formule_str(ws_formules[cellule])
        valeur_calculee = ws_valeurs[cellule].value

        if formule_obligatoire and formule is None:
            details.append(f"{cellule} : pas de formule (0 pt)")
            continue
        if not _contient_fonction(formule, fonctions_attendues):
            detectee = _nom_fonction_detectee(formule)
            attendues = "/".join(fonctions_attendues)
            if detectee:
                details.append(f"{cellule} : fonction {detectee} utilisee au lieu de {attendues} (0 pt)")
            else:
                details.append(f"{cellule} : aucune fonction {attendues} reconnue dans la formule (0 pt)")
            continue
        ref = valeurs_reference.get(cellule, valeur_reference)
        if not _valeur_proche(valeur_calculee, ref, tolerance):
            details.append(f"{cellule} : resultat incorrect (0 pt)")
            continue

        points_obtenus += points_par_cellule
        details.append(f"{cellule} : OK (+{points_par_cellule:.2f} pt)")

    return points_max, points_obtenus, details


def _corriger_formule_multiplication(ws_formules, ws_valeurs, critere):
    cellules = critere["cellules"]
    operandes = critere["operandes"]
    formule_obligatoire = critere.get("formule_obligatoire", True)
    valeurs_reference = critere.get("valeurs_reference", {})
    tolerance = critere.get("tolerance", 0.01)

    points_max = critere.get("points_par_cellule", 0) * len(cellules) if critere.get("points_par_cellule") else critere.get("points", len(cellules))
    points_par_cellule = points_max / len(cellules)
    points_obtenus = 0.0
    details = []
    col_a, col_b = operandes

    for cellule in cellules:
        formule = _formule_str(ws_formules[cellule])
        valeur_calculee = ws_valeurs[cellule].value
        ligne = re.search(r"\d+", cellule).group()

        if formule_obligatoire and formule is None:
            details.append(f"{cellule} : pas de formule (0 pt)")
            continue

        motif = rf"{col_a}{ligne}\s*\*\s*{col_b}{ligne}|{col_b}{ligne}\s*\*\s*{col_a}{ligne}"
        if not re.search(motif, formule):
            details.append(f"{cellule} : formule de multiplication incorrecte (0 pt)")
            continue

        ref = valeurs_reference.get(cellule)
        if not _valeur_proche(valeur_calculee, ref, tolerance):
            details.append(f"{cellule} : resultat incorrect (0 pt)")
            continue

        points_obtenus += points_par_cellule
        details.append(f"{cellule} : OK (+{points_par_cellule:.2f} pt)")

    return points_max, points_obtenus, details


def _corriger_formule_operateurs_mixtes(ws_formules, ws_valeurs, critere):
    """Verifie qu'une formule combine plusieurs operateurs arithmetiques ET donne le bon resultat."""
    cellules = critere["cellules"]
    nb_operateurs_min = critere.get("nb_operateurs_min", 2)
    valeurs_reference = critere.get("valeurs_reference", {})
    tolerance = critere.get("tolerance", 0.02)
    formule_obligatoire = critere.get("formule_obligatoire", True)

    points_max = critere.get("points", len(cellules))
    points_par_cellule = points_max / len(cellules)
    points_obtenus = 0.0
    details = []

    for cellule in cellules:
        formule = _formule_str(ws_formules[cellule])
        valeur_calculee = ws_valeurs[cellule].value

        if formule_obligatoire and formule is None:
            details.append(f"{cellule} : pas de formule (0 pt)")
            continue

        operateurs = set(re.findall(r"[+\-*/]", formule))
        if len(operateurs) < nb_operateurs_min:
            trouves = ", ".join(sorted(operateurs)) if operateurs else "aucun"
            details.append(
                f"{cellule} : un seul operateur ({trouves}) — combinez au moins {nb_operateurs_min} operateurs differents (0 pt)"
            )
            continue

        ref = valeurs_reference.get(cellule)
        if not _valeur_proche(valeur_calculee, ref, tolerance):
            details.append(f"{cellule} : resultat incorrect (0 pt)")
            continue

        points_obtenus += points_par_cellule
        details.append(f"{cellule} : OK (+{points_par_cellule:.2f} pt)")

    return points_max, points_obtenus, details


def _corriger_recopie_formule(ws_formules, ws_valeurs, critere):
    """Verifie une formule recopiee combinant reference relative (ligne courante)
    et reference absolue ($) vers une cellule fixe (ex: =B4*$B$1)."""
    cellules = critere["cellules"]
    cellule_absolue = critere["cellule_absolue"]
    colonne_relative = critere["colonne_relative"]
    valeurs_reference = critere.get("valeurs_reference", {})
    tolerance = critere.get("tolerance", 0.01)
    formule_obligatoire = critere.get("formule_obligatoire", True)

    col_abs = re.match(r"([A-Z]+)(\d+)", cellule_absolue).group(1)
    row_abs = re.match(r"([A-Z]+)(\d+)", cellule_absolue).group(2)
    motif_absolu = rf"\${col_abs}\${row_abs}"

    points_max = critere.get("points", len(cellules))
    points_par_cellule = points_max / len(cellules)
    points_obtenus = 0.0
    details = []

    for cellule in cellules:
        formule = _formule_str(ws_formules[cellule])
        valeur_calculee = ws_valeurs[cellule].value
        ligne = re.search(r"\d+", cellule).group()

        if formule_obligatoire and formule is None:
            details.append(f"{cellule} : pas de formule (0 pt)")
            continue
        if not re.search(motif_absolu, formule):
            details.append(f"{cellule} : reference absolue ${col_abs}${row_abs} manquante (0 pt)")
            continue
        motif_relatif_ancre = rf"\${colonne_relative}\$?{ligne}|{colonne_relative}\${ligne}"
        motif_relatif_plain = rf"(?<!\$){colonne_relative}{ligne}\b"
        if re.search(motif_relatif_ancre, formule) or not re.search(motif_relatif_plain, formule):
            details.append(f"{cellule} : reference relative {colonne_relative}{ligne} incorrecte ou ancree (0 pt)")
            continue
        ref = valeurs_reference.get(cellule)
        if not _valeur_proche(valeur_calculee, ref, tolerance):
            details.append(f"{cellule} : resultat incorrect (0 pt)")
            continue

        points_obtenus += points_par_cellule
        details.append(f"{cellule} : OK (+{points_par_cellule:.2f} pt)")

    return points_max, points_obtenus, details


def _a_une_couleur_de_fond(cell):
    fill = cell.fill
    return bool(fill and fill.fill_type == "solid" and fill.fgColor and fill.fgColor.rgb not in (None, "00000000"))


def _corriger_mise_en_forme_cellule(ws_valeurs, critere):
    """Verifie qu'une plage (typiquement une ligne d'en-tete) est en gras avec une couleur de fond."""
    points_max = critere.get("points", 1)
    plage = critere["plage"]

    cellules = [c for ligne in ws_valeurs[plage] for c in ligne]
    gras_ok = all(c.font and c.font.bold for c in cellules)
    couleur_ok = all(_a_une_couleur_de_fond(c) for c in cellules)
    ok = gras_ok and couleur_ok

    if ok:
        details = ["Gras et couleur de fond appliques sur la plage"]
    else:
        manques = []
        if not gras_ok:
            manques.append("gras manquant sur au moins une cellule")
        if not couleur_ok:
            manques.append("couleur de fond manquante sur au moins une cellule")
        details = [", ".join(manques)]

    return points_max, (points_max if ok else 0), details


def _corriger_format_nombre(ws_valeurs, critere):
    """Verifie qu'un format d'affichage (ex: pourcentage) est applique sans changer la valeur reelle."""
    cellules = critere["cellules"]
    format_attendu = critere.get("format_attendu", "%")
    valeurs_reference = critere.get("valeurs_reference", {})
    tolerance = critere.get("tolerance", 0.01)

    points_max = critere.get("points", len(cellules))
    points_par_cellule = points_max / len(cellules)
    points_obtenus = 0.0
    details = []

    for cellule in cellules:
        format_cellule = ws_valeurs[cellule].number_format
        valeur_calculee = ws_valeurs[cellule].value
        ref = valeurs_reference.get(cellule)

        if format_attendu not in format_cellule:
            details.append(f"{cellule} : format d'affichage attendu absent (0 pt)")
            continue
        if not _valeur_proche(valeur_calculee, ref, tolerance):
            details.append(f"{cellule} : la valeur reelle a ete modifiee, seul l'affichage doit changer (0 pt)")
            continue

        points_obtenus += points_par_cellule
        details.append(f"{cellule} : OK (+{points_par_cellule:.2f} pt)")

    return points_max, points_obtenus, details


def _comparer_lignes(ligne_a, ligne_b, colonnes):
    for col in colonnes:
        va, vb = ligne_a.get(col["colonne"]), ligne_b.get(col["colonne"])
        if va == vb:
            continue
        if va is None:
            return -1
        if vb is None:
            return 1
        try:
            resultat = -1 if va < vb else 1
        except TypeError:
            resultat = -1 if str(va) < str(vb) else 1
        return -resultat if col.get("ordre") == "desc" else resultat
    return 0


def _corriger_tri(ws_valeurs, critere):
    plage = critere["plage"]
    colonnes = critere["colonnes"]
    points_max = critere.get("points", 1)

    lignes_actuelles = [
        {get_column_letter(c.column): c.value for c in ligne} for ligne in ws_valeurs[plage]
    ]
    lignes_attendues = sorted(lignes_actuelles, key=cmp_to_key(lambda a, b: _comparer_lignes(a, b, colonnes)))

    ok = lignes_actuelles == lignes_attendues
    details = ["Ordre de tri correct" if ok else "Ordre de tri incorrect"]
    return points_max, (points_max if ok else 0), details


def _corriger_filtre_auto(ws_valeurs, critere):
    points_max = critere.get("points", 1)
    plage_attendue = critere.get("plage_attendue")
    actif = bool(ws_valeurs.auto_filter.ref)
    ok = actif and (plage_attendue is None or ws_valeurs.auto_filter.ref == plage_attendue)
    details = ["Filtre automatique actif" if ok else "Filtre automatique absent ou incorrect"]
    return points_max, (points_max if ok else 0), details


def _corriger_mise_en_forme_conditionnelle(ws_valeurs, critere):
    points_max = critere.get("points", 1)
    plage = critere.get("plage")
    regles = list(ws_valeurs.conditional_formatting)
    if plage:
        ok = any(plage in str(r.sqref) or str(r.sqref) in plage for r in regles)
    else:
        ok = len(regles) > 0
    details = ["Mise en forme conditionnelle detectee" if ok else "Mise en forme conditionnelle absente"]
    return points_max, (points_max if ok else 0), details


def _extraire_texte_titre(chart):
    """Extrait le texte brut du titre d'un graphique openpyxl. Retourne '' si absent ou vide."""
    try:
        title = chart.title
        if title is None:
            return ""
        tx = title.tx
        if tx is None:
            return ""
        rich = tx.rich
        if rich is None:
            return ""
        texte = ""
        for p in rich.p:
            for r in p.r:
                texte += (r.t or "")
        return texte.strip()
    except Exception:
        return ""


def _corriger_graphique(ws_valeurs, critere):
    from openpyxl.chart import BarChart, LineChart, PieChart

    points_max = critere.get("points", 1)
    type_impose = critere.get("type_impose")     # "histogramme" | "courbes" | "secteurs" | None
    titre_oblige = critere.get("titre_oblige", False)
    tendance_obligee = critere.get("tendance", False)

    nb_criteres = 1 + (1 if type_impose else 0) + (1 if titre_oblige else 0) + (1 if tendance_obligee else 0)
    pts_par = points_max / nb_criteres

    charts = ws_valeurs._charts
    if not charts:
        return points_max, 0, ["Aucun graphique detecte (0 pt)"]

    chart = charts[0]
    points_obtenus = pts_par
    details = [f"Graphique detecte : +{pts_par:.2f} pt"]

    TYPE_MAP = {"histogramme": BarChart, "courbes": LineChart, "secteurs": PieChart}
    NOMS_FR = {"BarChart": "histogramme", "LineChart": "courbes", "PieChart": "secteurs"}

    if type_impose:
        cls = TYPE_MAP.get(type_impose)
        if cls and isinstance(chart, cls):
            points_obtenus += pts_par
            details.append(f"Type correct ({type_impose}) : +{pts_par:.2f} pt")
        else:
            trouve = NOMS_FR.get(type(chart).__name__, type(chart).__name__)
            details.append(f"Type incorrect : {trouve} trouve, {type_impose} attendu (0 pt)")

    if titre_oblige:
        texte_titre = _extraire_texte_titre(chart)
        mots_cles = critere.get("mots_cles_titre", ["ventes"])
        titre_attendu = critere.get("titre_attendu", "")
        titre_normalise = texte_titre.lower()
        mots_trouves = [m for m in mots_cles if m.lower() in titre_normalise]

        if not texte_titre:
            details.append(
                f"Titre absent ou vide — titre attendu : \"{titre_attendu}\" (0 pt)"
            )
        elif len(mots_trouves) < len(mots_cles):
            mots_manquants = [m for m in mots_cles if m.lower() not in titre_normalise]
            details.append(
                f"Titre present (\"{texte_titre}\") mais incomplet — "
                f"element(s) manquant(s) : {', '.join(mots_manquants)} (0 pt)"
            )
        else:
            points_obtenus += pts_par
            details.append(f"Titre correct (\"{texte_titre}\") : +{pts_par:.2f} pt")

    if tendance_obligee:
        trendline = None
        try:
            if chart.series and chart.series[0].trendline is not None:
                trendline = chart.series[0].trendline
        except Exception:
            pass
        if trendline is None:
            details.append("Courbe de tendance absente (0 pt)")
        elif trendline.trendlineType != "linear":
            noms_type = {
                "log": "logarithmique", "exp": "exponentielle", "poly": "polynomiale",
                "power": "puissance", "movingAvg": "moyenne mobile",
            }
            nom_trouve = noms_type.get(trendline.trendlineType, trendline.trendlineType)
            details.append(f"Courbe de tendance incorrecte : {nom_trouve} trouvee, lineaire attendue (0 pt)")
        else:
            points_obtenus += pts_par
            details.append(f"Courbe de tendance lineaire presente : +{pts_par:.2f} pt")

    return points_max, points_obtenus, details


def _corriger_tcd(chemin_fichier, critere):
    points_max = critere.get("points", 1)
    with zipfile.ZipFile(chemin_fichier) as archive:
        ok = any(nom.startswith("xl/pivotTables/") for nom in archive.namelist())
    details = ["Tableau croise dynamique detecte" if ok else "Aucun tableau croise dynamique detecte"]
    return points_max, (points_max if ok else 0), details


def _corriger_texte_motcles(ws_valeurs, critere):
    points_max = critere.get("points", 1)
    cellule = critere["cellule"]
    mots_cles = [m.lower() for m in critere["mots_cles"]]
    mode = critere.get("mode", "un_de")

    texte = str(ws_valeurs[cellule].value or "").lower()
    trouves = [m for m in mots_cles if m in texte]
    ok = len(trouves) == len(mots_cles) if mode == "tous" else len(trouves) > 0

    details = [f"Mots-cles trouves : {trouves}" if trouves else "Aucun mot-cle trouve"]
    return points_max, (points_max if ok else 0), details


_DISPATCH = {
    "formule_fonction": lambda wf, wv, cf, critere: _corriger_formule_fonction(wf, wv, critere),
    "formule_multiplication": lambda wf, wv, cf, critere: _corriger_formule_multiplication(wf, wv, critere),
    "formule_operateurs_mixtes": lambda wf, wv, cf, critere: _corriger_formule_operateurs_mixtes(wf, wv, critere),
    "tri": lambda wf, wv, cf, critere: _corriger_tri(wv, critere),
    "filtre_auto": lambda wf, wv, cf, critere: _corriger_filtre_auto(wv, critere),
    "mise_en_forme_conditionnelle": lambda wf, wv, cf, critere: _corriger_mise_en_forme_conditionnelle(wv, critere),
    "graphique": lambda wf, wv, cf, critere: _corriger_graphique(wv, critere),
    "tcd": lambda wf, wv, cf, critere: _corriger_tcd(cf, critere),
    "texte_motcles": lambda wf, wv, cf, critere: _corriger_texte_motcles(wv, critere),
    "recopie_formule": lambda wf, wv, cf, critere: _corriger_recopie_formule(wf, wv, critere),
    "mise_en_forme_cellule": lambda wf, wv, cf, critere: _corriger_mise_en_forme_cellule(wv, critere),
    "format_nombre": lambda wf, wv, cf, critere: _corriger_format_nombre(wv, critere),
}


def corriger_critere(wb_formules, wb_valeurs, chemin_fichier, exercice, critere):
    type_critere = critere["type"]
    if type_critere not in _DISPATCH:
        raise ValueError(f"Type de critere inconnu : {type_critere}")

    feuille = critere.get("feuille", exercice.get("feuille"))
    if feuille not in wb_valeurs.sheetnames:
        return critere.get("points", critere.get("points_par_cellule", 1)), 0, [f"Feuille '{feuille}' absente du classeur"]

    ws_formules = wb_formules[feuille]
    ws_valeurs = wb_valeurs[feuille]
    return _DISPATCH[type_critere](ws_formules, ws_valeurs, chemin_fichier, critere)


def corriger_copie(chemin_fichier, config):
    """Corrige une copie etudiante selon une config d'epreuve et renvoie le resultat structure."""
    wb_formules = openpyxl.load_workbook(chemin_fichier, data_only=False)
    wb_valeurs = openpyxl.load_workbook(chemin_fichier, data_only=True)

    identite = _identite_depuis_feuille(wb_valeurs) or extraire_identite(chemin_fichier)

    resultat = {
        "identite": identite,
        "session": config.get("session"),
        "annee": config.get("annee"),
        "module": config.get("module"),
        "variante": config.get("variante"),
        "watermark": verifier_watermark(wb_formules, config),
        "exercices": [],
        "points_obtenus": 0.0,
        "points_max": 0.0,
    }

    for exercice in config["exercices"]:
        ex_resultat = {
            "id": exercice["id"],
            "titre": exercice["titre"],
            "criteres": [],
            "points_obtenus": 0.0,
            "points_max": 0.0,
        }
        for critere in exercice["criteres"]:
            points_max, points_obtenus, details = corriger_critere(
                wb_formules, wb_valeurs, chemin_fichier, exercice, critere
            )
            statut = STATUT_REUSSI if points_obtenus >= points_max - 1e-6 else STATUT_ECHOUE
            ex_resultat["criteres"].append({
                "id": critere["id"],
                "competence": critere.get("competence", critere.get("description", "")),
                "description": critere.get("description", ""),
                "points_max": points_max,
                "points_obtenus": points_obtenus,
                "statut": statut,
                "details": details,
            })
            ex_resultat["points_obtenus"] += points_obtenus
            ex_resultat["points_max"] += points_max

        resultat["exercices"].append(ex_resultat)
        resultat["points_obtenus"] += ex_resultat["points_obtenus"]
        resultat["points_max"] += ex_resultat["points_max"]

    resultat["bareme_total"] = config.get("bareme_total", resultat["points_max"])
    return resultat


def resumer_echecs(details, max_cellules=4):
    """Condense les details d'un critere en un resume avec references de cellules.

    Conserve les coordonnees de cellule pour que l'etudiant sache exactement
    ou corriger (ex: 'D5 : formule incorrecte ; D6 : resultat incorrect').
    """
    if not details:
        return ""
    echecs = [d for d in details if " : OK" not in d]
    if not echecs:
        return ""

    parties = [re.sub(r"\s*\(0 pt\)$", "", d) for d in echecs[:max_cellules]]
    if len(echecs) > max_cellules:
        parties.append(f"... et {len(echecs) - max_cellules} autre(s) cellule(s)")
    return " ; ".join(parties)


def resultats_vers_lignes_csv(resultat):
    lignes = []
    for exercice in resultat["exercices"]:
        for critere in exercice["criteres"]:
            lignes.append({
                "nom": resultat["identite"]["nom"],
                "prenom": resultat["identite"]["prenom"],
                "session": resultat["session"],
                "annee": resultat["annee"],
                "module": resultat["module"],
                "exercice": exercice["titre"],
                "critere": critere["description"],
                "points_obtenus": critere["points_obtenus"],
                "points_max": critere["points_max"],
                "statut": critere["statut"],
                "details": " | ".join(critere["details"]),
            })
    return lignes


def exporter_csv(resultats, chemin_sortie):
    lignes = []
    for resultat in resultats:
        lignes.extend(resultats_vers_lignes_csv(resultat))
    if not lignes:
        return
    with open(chemin_sortie, "w", newline="", encoding="utf-8") as fichier:
        writer = csv.DictWriter(fichier, fieldnames=lignes[0].keys())
        writer.writeheader()
        writer.writerows(lignes)
