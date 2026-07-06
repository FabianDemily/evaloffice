"""Tracabilite des fichiers generes : watermark UUID + hash des zones non editables.

Concu pour etre reutilise par les futurs modules Word/IA (phase 2) : le registre
et le calcul de hash sont generiques (independants du format de fichier) ; seule
l'incrustation/lecture du watermark dans le document est specifique au format
(xlsx aujourd'hui via openpyxl, docx plus tard via python-docx).

Ce mecanisme NE detecte PAS l'usage d'une IA. Il detecte :
- une substitution complete du fichier distribue (identifiant absent/different) ;
- une falsification des donnees source (hash des zones figees different).
"""

import hashlib
import uuid as uuid_lib

from openpyxl.utils import get_column_letter, range_boundaries

PREFIXE_WATERMARK = "evaloffice-id:"
FEUILLE_GARDE = "0 - Consignes"


def generer_uuid():
    return str(uuid_lib.uuid4())


def definir_watermark_xlsx(wb, identifiant):
    """Incruste l'identifiant dans les proprietes du document (visible si on regarde
    les proprietes du fichier, mais ce n'est pas une information sensible a cacher)."""
    wb.properties.keywords = f"{PREFIXE_WATERMARK}{identifiant}"


def lire_watermark_xlsx(wb):
    mots_cles = wb.properties.keywords or ""
    if mots_cles.startswith(PREFIXE_WATERMARK):
        return mots_cles[len(PREFIXE_WATERMARK):]
    return None


def _coords_plage(plage):
    min_col, min_row, max_col, max_row = range_boundaries(plage)
    return {
        f"{get_column_letter(c)}{r}"
        for r in range(min_row, max_row + 1)
        for c in range(min_col, max_col + 1)
    }


def cellules_editables(config):
    """Construit {feuille: {coordonnees}} des cellules que l'etudiant doit remplir,
    a partir des criteres de la config. Le reste des cellules est considere figé
    (donnees source, en-tetes) et sert au calcul du hash d'integrite."""
    editables = {}
    for exercice in config.get("exercices", []):
        feuille = exercice.get("feuille")
        if not feuille:
            continue
        coords = editables.setdefault(feuille, set())
        for critere in exercice["criteres"]:
            if "cellules" in critere:
                coords.update(critere["cellules"])
            if "cellule" in critere:
                coords.add(critere["cellule"])
            if "plage" in critere:
                coords.update(_coords_plage(critere["plage"]))
    return editables


def hash_zones_figees(wb, config):
    """SHA-256 du contenu des cellules NON editables (donnees source, en-tetes),
    pour detecter une falsification des donnees de base (ex: changer les prix
    pour que le resultat attendu corresponde sans calcul reel)."""
    editables = cellules_editables(config)
    empreinte = hashlib.sha256()
    for feuille_nom in sorted(wb.sheetnames):
        if feuille_nom == FEUILLE_GARDE:
            continue
        ws = wb[feuille_nom]
        coords_editables = editables.get(feuille_nom, set())
        for ligne in ws.iter_rows():
            for cell in ligne:
                if cell.value is None or cell.coordinate in coords_editables:
                    continue
                valeur = cell.value
                # Normalise les flottants entiers (24.0) car le cycle sauvegarde/relecture
                # xlsx peut les convertir en int (24) sans que ce soit une vraie modification.
                if isinstance(valeur, float) and valeur.is_integer():
                    valeur = int(valeur)
                empreinte.update(f"{feuille_nom}|{cell.coordinate}|{valeur}".encode("utf-8"))
    return empreinte.hexdigest()


def verifier_watermark(wb, config):
    """Compare le watermark/hash d'un fichier soumis a ceux attendus par la config.

    Renvoie None si la config n'a pas de watermark active. Sinon un dict avec
    des indicateurs de vigilance (jamais une preuve, toujours a verifier manuellement).
    """
    attendu = config.get("watermark")
    if not attendu:
        return None

    id_trouve = lire_watermark_xlsx(wb)
    id_ok = id_trouve is not None and id_trouve == attendu.get("id")

    hash_trouve = hash_zones_figees(wb, config)
    hash_ok = hash_trouve == attendu.get("hash_zones_figees")

    messages = []
    if not id_ok:
        messages.append(
            "Identifiant de tracabilite absent ou different : le fichier soumis ne "
            "correspond pas au fichier distribue (substitution possible — a verifier manuellement)."
        )
    if not hash_ok:
        messages.append(
            "Les donnees source (hors cellules a completer) different du fichier distribue "
            "(falsification possible des donnees de base — a verifier manuellement)."
        )

    return {"id_ok": id_ok, "hash_ok": hash_ok, "ok": id_ok and hash_ok, "messages": messages}
