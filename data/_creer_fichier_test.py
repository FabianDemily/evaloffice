"""Script ponctuel pour generer un fichier .xlsx de test correspondant a la config exemple."""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font

wb = Workbook()
wb.remove(wb.active)

# Ex1 - Formules
ws1 = wb.create_sheet("Ex1 - Formules")
ws1.append(["Titre", "Prix", "Quantite", "CA", "Stock OK"])
donnees = [
    ("Livre A", 10, 5),
    ("Livre B", 15, 8),
    ("Livre C", 20, 3),
    ("Livre D", 12, 7),
    ("Livre E", 16, 9),
]
for i, (titre, prix, qte) in enumerate(donnees, start=5):
    ws1[f"A{i}"] = titre
    ws1[f"B{i}"] = prix
    ws1[f"C{i}"] = qte
    ws1[f"D{i}"] = f"=B{i}*C{i}"
    ws1[f"E{i}"] = f"=SI(C{i}>5,\"OK\",\"BAS\")"
ws1["C10"] = "=SOMME(C5:C9)"
ws1["B11"] = "=MOYENNE(B5:B9)"

# Ex2 - Tri Filtres
ws2 = wb.create_sheet("Ex2 - Tri Filtres")
ws2.append(["Titre", "Categorie", "Prix", "Stock"])
livres_tries = [
    ("Atlas", "Geographie", 18, 4),
    ("Biologie 101", "Sciences", 22, 6),
    ("Chimie facile", "Sciences", 19, 9),
    ("Dictees CE2", "Scolaire", 9, 12),
    ("Encyclopedie", "Sciences", 35, 2),
    ("Francais facile", "Scolaire", 11, 15),
    ("Geographie monde", "Geographie", 21, 7),
    ("Histoire ancienne", "Histoire", 17, 5),
    ("Initiation calcul", "Scolaire", 8, 20),
    ("Jardins du monde", "Geographie", 14, 3),
    ("Kit experiences", "Sciences", 29, 6),
    ("Lecture CP", "Scolaire", 7, 18),
    ("Mathematiques 6e", "Scolaire", 13, 9),
    ("Nature et vivant", "Sciences", 16, 11),
]
for i, ligne in enumerate(livres_tries, start=2):
    for j, val in enumerate(ligne):
        ws2.cell(row=i, column=j + 1, value=val)
ws2.auto_filter.ref = "A1:D15"

# Deja triee : Categorie (asc) puis Prix (desc) -- simule une copie correcte
livres_multicritere = [
    ("Cartes anciennes", "Geographie", 27, 3),
    ("Atlas Pro", "Geographie", 18, 4),
    ("Jungle amazonienne", "Geographie", 16, 5),
    ("Fleuves du monde", "Geographie", 15, 9),
    ("Lacs et montagnes", "Geographie", 13, 10),
    ("Iles tropicales", "Geographie", 12, 11),
    ("Empires perdus", "Histoire", 24, 4),
    ("Kingdoms anciens", "Histoire", 21, 3),
    ("Guerre et paix", "Histoire", 20, 6),
    ("Histoire", "Sciences", 30, 2),
    ("Geologie", "Sciences", 25, 5),
    ("Botanique", "Sciences", 22, 6),
    ("Decouverte espace", "Sciences", 19, 8),
    ("Hydrologie", "Sciences", 17, 7),
]
for i, ligne in enumerate(livres_multicritere, start=18):
    for j, val in enumerate(ligne):
        ws2.cell(row=i, column=j + 1, value=val)

# Ex3 - Mise en forme
ws3 = wb.create_sheet("Ex3 - Mise en forme")
ws3.append(["Categorie", "Ventes"])
for i, (cat, ventes) in enumerate([("Romans", 120), ("BD", 80), ("Sciences", 95)], start=2):
    ws3[f"A{i}"] = cat
    ws3[f"B{i}"] = ventes
ws3.conditional_formatting.add(
    "B2:B20", CellIsRule(operator="greaterThan", formula=["100"], fill=None)
)
chart = BarChart()
data = Reference(ws3, min_col=2, min_row=1, max_row=4)
chart.add_data(data, titles_from_data=True)
ws3.add_chart(chart, "D2")

# Ex4 - TCD (pas de vrai TCD cree pour ce test -> volontairement absent)
ws4 = wb.create_sheet("Ex4 - TCD")
ws4["A1"] = "Pas de TCD dans ce test"

# Ex5 - Synthese
ws5 = wb.create_sheet("Ex5 - Synthese")
ws5["A1"] = "Explication"
ws5["B2"] = "J'ai trie les livres par ordre croissant alphabetique."

wb.save("templates/test_etudiant.xlsx")
print("Fichier de test cree : templates/test_etudiant.xlsx")
