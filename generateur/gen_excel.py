"""Generation d'epreuves Excel (fichier etudiant + corrige) et de leur config JSON.

Le contenu genere depend entierement du catalogue de competences
(data/competences.json) et des competences actives passees en parametre.
Aucune valeur de reference ou de bareme n'est codee en dur : les references
(sommes, moyennes, seuils) sont calculees a partir des donnees generees, et
les points viennent du catalogue ou des surcharges fournies par l'appelant.
"""

import json
import random
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.trendline import Trendline
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

TYPES_GRAPHIQUE = {
    "histogramme": BarChart,   # colonnes verticales (BarChart barDir="col" par defaut)
    "courbes": LineChart,
    "secteurs": PieChart,
}

NOMS_FR_GRAPHIQUE = {
    "histogramme": "histogramme (colonnes verticales)",
    "courbes": "courbes / lignes",
    "secteurs": "secteurs (camembert)",
    "BarChart": "histogramme",
    "LineChart": "courbes",
    "PieChart": "secteurs",
}

from integrite import definir_watermark_xlsx, generer_uuid, hash_zones_figees

RACINE = Path(__file__).parent.parent
THEMES_PATH = RACINE / "data" / "themes.json"
COMPETENCES_PATH = RACINE / "data" / "competences.json"

# openpyxl ecrit les formules en anglais dans le XML xlsx.
# Excel/Calc les traduit ensuite selon la langue de l'interface.
FONCTIONS_EN = {
    "SOMME": "SUM", "MOYENNE": "AVERAGE", "MAX": "MAX", "MIN": "MIN",
    "NB": "COUNT", "NB.SI": "COUNTIF", "SI": "IF", "ET": "AND", "OU": "OR",
}


def charger_themes():
    return json.loads(THEMES_PATH.read_text(encoding="utf-8"))


def charger_competences():
    return json.loads(COMPETENCES_PATH.read_text(encoding="utf-8"))


def sauver_competences(competences):
    COMPETENCES_PATH.write_text(json.dumps(competences, indent=2, ensure_ascii=False), encoding="utf-8")


def _construire_donnees(theme, nb_lignes, rng):
    items = rng.sample(theme["items"], k=min(nb_lignes, len(theme["items"])))
    lignes = []
    for nom in items:
        lignes.append({
            "nom": nom,
            "categorie": rng.choice(theme["categories"]),
            "prix": round(rng.uniform(5, 35), 2),
            "quantite": rng.randint(2, 15),
            "taux_reussite": round(rng.uniform(0.5, 0.99), 2),
        })
    return lignes


def _trier(donnees, cles):
    resultat = list(donnees)
    for cle, ordre in reversed(cles):
        resultat = sorted(resultat, key=lambda d: d[cle], reverse=(ordre == "desc"))
    return resultat


def _points(competences_par_id, comp_id):
    return competences_par_id[comp_id].get("points_defaut", 1)


def _formater_consigne(comp, **kwargs):
    gabarit = comp.get("consigne", "")
    try:
        return gabarit.format(**kwargs)
    except (KeyError, IndexError):
        return gabarit


def _ecrire_consignes(ws, consignes, cellule="H1"):
    """Ecrit les consignes de l'exercice dans une cellule en marge du tableau,
    sans perturber la mise en page des donnees (colonnes A-D/E utilisees)."""
    if not consignes:
        return
    texte = "Consignes :\n" + "\n".join(f"- {c}" for c in consignes)
    cell = ws[cellule]
    cell.value = texte
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(bold=False)
    ws.column_dimensions[cell.column_letter].width = 55


def _calculer_agregat(fonction, donnees, colonne):
    cle = {"B": "prix", "C": "quantite"}.get(colonne)
    if cle is None:
        return None
    valeurs = [d[cle] for d in donnees]
    if not valeurs:
        return None
    if fonction == "SOMME":
        return round(sum(valeurs), 2)
    if fonction == "MOYENNE":
        return round(sum(valeurs) / len(valeurs), 2)
    if fonction == "MAX":
        return max(valeurs)
    if fonction == "MIN":
        return min(valeurs)
    if fonction == "NB":
        return len(valeurs)
    return None  # fonction inconnue/personnalisee : pas de verification de resultat, juste de syntaxe


def _construire_groupe_formules(wb_etudiant, wb_corrige, theme, donnees, competences_par_id, actives):
    """Construit l'exercice 'Formules de base'.

    Generique : toute competence du groupe 'formules' de type formule_fonction
    (autre que formule_si, qui est conditionnelle par ligne) devient une ligne
    d'agregat, ce qui permet d'ajouter une nouvelle fonction depuis le catalogue
    sans modifier ce module.
    """
    comp_formules = [c for cid, c in competences_par_id.items() if cid in actives and c.get("groupe") == "formules"]
    if not comp_formules:
        return None

    feuille = "Ex1 - Formules"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)

    comp_multiplication = next((c for c in comp_formules if c["type"] == "formule_multiplication"), None)
    comp_op_mixtes = next((c for c in comp_formules if c["type"] == "formule_operateurs_mixtes"), None)
    comp_si = next((c for c in comp_formules if c["id"] == "formule_si"), None)
    comp_agregats = [c for c in comp_formules if c["type"] == "formule_fonction" and c is not comp_si]

    # Attribution dynamique des colonnes apres A (item), B (prix), C (quantite)
    prochaine = ord("D")
    col_ca = col_marge = col_op = col_si = None
    if comp_multiplication:
        col_ca = chr(prochaine); prochaine += 1
    if comp_op_mixtes:
        col_marge = chr(prochaine); prochaine += 1
        col_op = chr(prochaine); prochaine += 1
    if comp_si:
        col_si = chr(prochaine); prochaine += 1

    entetes = [theme["colonne_item"], "Prix", "Quantite"]
    if col_ca:     entetes.append("Chiffre d'affaires")
    if col_marge:  entetes.append("Marge unitaire")
    if col_op:     entetes.append("CA total (Prix + Marge) x Qte")
    if col_si:     entetes.append("Stock suffisant (>5)")
    for ws in (ws_e, ws_c):
        ws.append(entetes)

    debut = 2
    valeurs_reference_op = {}
    for i, ligne in enumerate(donnees):
        r = debut + i
        marge = round(ligne["prix"] * 0.2, 2)
        for ws in (ws_e, ws_c):
            ws[f"A{r}"] = ligne["nom"]
            ws[f"B{r}"] = ligne["prix"]
            ws[f"C{r}"] = ligne["quantite"]
        if col_ca:
            ws_c[f"{col_ca}{r}"] = f"=B{r}*C{r}"
        if col_marge:
            for ws in (ws_e, ws_c):
                ws[f"{col_marge}{r}"] = marge
        if col_op:
            ws_c[f"{col_op}{r}"] = f"=({col_marge}{r}+B{r})*C{r}"
            valeurs_reference_op[f"{col_op}{r}"] = round((ligne["prix"] + marge) * ligne["quantite"], 2)
        if col_si:
            ws_c[f"{col_si}{r}"] = f"=IF(C{r}>5,\"OK\",\"BAS\")"
    fin = debut + len(donnees) - 1

    criteres = []
    consignes = []
    if comp_multiplication:
        cellules_ca = [f"{col_ca}{debut + i}" for i in range(len(donnees))]
        criteres.append({
            "id": comp_multiplication["id"], "competence": comp_multiplication["label"],
            "description": "Chiffre d'affaires : formule Prix x Quantite",
            "type": "formule_multiplication", "cellules": cellules_ca, "operandes": ["B", "C"],
            "points": comp_multiplication.get("points_defaut", 1), "formule_obligatoire": True,
        })
        consignes.append(_formater_consigne(
            comp_multiplication, colonne_resultat=col_ca,
            cellules=f"{col_ca}{debut}:{col_ca}{fin}", libelle="chiffre d'affaires (Prix x Quantite)"
        ))

    if comp_op_mixtes:
        cellules_op = [f"{col_op}{debut + i}" for i in range(len(donnees))]
        criteres.append({
            "id": comp_op_mixtes["id"], "competence": comp_op_mixtes["label"],
            "description": f"CA total = (Prix + Marge) x Quantite — formule avec operateurs mixtes",
            "type": "formule_operateurs_mixtes", "cellules": cellules_op,
            "valeurs_reference": valeurs_reference_op, "nb_operateurs_min": 2,
            "tolerance": 0.02, "points": comp_op_mixtes.get("points_defaut", 3), "formule_obligatoire": True,
        })
        consignes.append(_formater_consigne(
            comp_op_mixtes, colonne_resultat=col_op,
            cellules=f"{col_op}{debut}:{col_op}{fin}",
            libelle=f"CA total en colonne {col_op} : (Prix + Marge) x Quantite"
        ))

    ligne_agregat = fin + 1
    for comp in comp_agregats:
        fonction = comp.get("fonction", "").upper()
        colonne = comp.get("colonne", "C")
        libelle = comp.get("libelle", comp.get("label", fonction))
        cell = f"{colonne}{ligne_agregat}"
        for ws in (ws_e, ws_c):
            ws[f"A{ligne_agregat}"] = f"{libelle} :"
        fonction_en = FONCTIONS_EN.get(fonction, fonction)
        ws_c[cell] = f"={fonction_en}({colonne}{debut}:{colonne}{fin})"
        critere = {
            "id": comp["id"], "competence": comp["label"], "description": f"{fonction} : {libelle}", "type": "formule_fonction",
            "cellules": [cell], "fonctions_attendues": [fonction], "tolerance": 0.01,
            "points": comp.get("points_defaut", 1), "formule_obligatoire": True,
        }
        valeur_ref = _calculer_agregat(fonction, donnees, colonne)
        if valeur_ref is not None:
            critere["valeur_reference"] = valeur_ref
        criteres.append(critere)
        consignes.append(_formater_consigne(comp, cellule=cell, fonction=fonction, libelle=libelle))
        ligne_agregat += 1

    if comp_si:
        cellules_si = [f"{col_si}{debut + i}" for i in range(len(donnees))]
        valeurs_reference_si = {
            f"{col_si}{debut + i}": ("OK" if ligne["quantite"] > 5 else "BAS")
            for i, ligne in enumerate(donnees)
        }
        criteres.append({
            "id": comp_si["id"], "competence": comp_si["label"],
            "description": "SI conditionnel sur le niveau de stock", "type": "formule_fonction",
            "cellules": cellules_si, "fonctions_attendues": ["SI"], "valeurs_reference": valeurs_reference_si,
            "points": comp_si.get("points_defaut", 1), "formule_obligatoire": True,
        })
        consignes.append(_formater_consigne(comp_si, cellules=f"{col_si}{debut}:{col_si}{fin}"))

    _ecrire_consignes(ws_e, consignes)

    return {"id": "ex1", "titre": "Formules de base", "feuille": feuille, "criteres": criteres, "consignes": consignes}


def _construire_groupe_tri(wb_etudiant, wb_corrige, theme, donnees_a, donnees_b, competences_par_id, actives, rng):
    pertinentes = [c for c in ("tri_simple", "tri_multicritere", "filtre_auto") if c in actives]
    if not pertinentes:
        return None

    feuille = "Ex2 - Tri Filtres"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    entetes = [theme["colonne_item"], "Categorie", "Prix", "Stock"]

    table_a = "tri_simple" in actives or "filtre_auto" in actives
    table_b = "tri_multicritere" in actives

    criteres = []
    consignes = []
    ligne_courante = 1
    fin_a = None

    if table_a:
        debut_a = ligne_courante + 1
        ws_e.append(entetes)
        ws_c.append(entetes)
        melangees = list(donnees_a)
        rng.shuffle(melangees)
        triees = _trier(donnees_a, [("nom", "asc")])
        for i, (le, lc) in enumerate(zip(melangees, triees)):
            r = debut_a + i
            for col, cle in zip("ABCD", ("nom", "categorie", "prix", "quantite")):
                ws_e[f"{col}{r}"] = le[cle]
                ws_c[f"{col}{r}"] = lc[cle]
        fin_a = debut_a + len(donnees_a) - 1
        ligne_courante = fin_a

        if "tri_simple" in actives:
            plage_a = f"A{debut_a}:D{fin_a}"
            criteres.append({
                "id": "tri_simple", "competence": competences_par_id["tri_simple"]["label"], "description": f"Tri croissant sur la colonne {theme['colonne_item']}",
                "type": "tri", "plage": plage_a,
                "colonnes": [{"colonne": "A", "ordre": "asc"}], "points": _points(competences_par_id, "tri_simple"),
            })
            consignes.append(_formater_consigne(
                competences_par_id["tri_simple"], plage=plage_a, colonne_item=theme["colonne_item"]
            ))
        if "filtre_auto" in actives:
            ws_c.auto_filter.ref = f"A1:D{fin_a}"
            criteres.append({
                "id": "filtre_auto", "competence": competences_par_id["filtre_auto"]["label"], "description": "Filtre automatique active sur le tableau",
                "type": "filtre_auto", "feuille": feuille, "points": _points(competences_par_id, "filtre_auto"),
            })
            consignes.append(_formater_consigne(competences_par_id["filtre_auto"]))

    if table_b:
        debut_b = ligne_courante + 3
        for ws in (ws_e, ws_c):
            ws[f"A{debut_b - 1}"] = entetes[0]
            ws[f"B{debut_b - 1}"] = entetes[1]
            ws[f"C{debut_b - 1}"] = entetes[2]
            ws[f"D{debut_b - 1}"] = entetes[3]
        melangees = list(donnees_b)
        rng.shuffle(melangees)
        triees = _trier(donnees_b, [("categorie", "asc"), ("prix", "desc")])
        for i, (le, lc) in enumerate(zip(melangees, triees)):
            r = debut_b + i
            for col, cle in zip("ABCD", ("nom", "categorie", "prix", "quantite")):
                ws_e[f"{col}{r}"] = le[cle]
                ws_c[f"{col}{r}"] = lc[cle]
        fin_b = debut_b + len(donnees_b) - 1

        plage_b = f"A{debut_b}:D{fin_b}"
        criteres.append({
            "id": "tri_multicritere", "competence": competences_par_id["tri_multicritere"]["label"], "description": "Tri multicritere : Categorie (asc) puis Prix (desc)",
            "type": "tri", "plage": plage_b,
            "colonnes": [{"colonne": "B", "ordre": "asc"}, {"colonne": "C", "ordre": "desc"}],
            "points": _points(competences_par_id, "tri_multicritere"),
        })
        consignes.append(_formater_consigne(competences_par_id["tri_multicritere"], plage=plage_b))

        if "filtre_auto" in actives and not table_a:
            ws_c.auto_filter.ref = f"A{debut_b - 1}:D{fin_b}"
            criteres.append({
                "id": "filtre_auto", "competence": competences_par_id["filtre_auto"]["label"], "description": "Filtre automatique active sur le tableau",
                "type": "filtre_auto", "feuille": feuille, "points": _points(competences_par_id, "filtre_auto"),
            })
            consignes.append(_formater_consigne(competences_par_id["filtre_auto"]))

    _ecrire_consignes(ws_e, consignes)

    return {"id": "ex2", "titre": "Tri et filtres", "feuille": feuille, "criteres": criteres, "consignes": consignes}


def _construire_groupe_visuel(wb_etudiant, wb_corrige, theme, competences_par_id, actives, rng, options_par_competence=None, contexte=""):
    pertinentes = [c for c in ("mise_en_forme_conditionnelle", "graphique") if c in actives]
    if not pertinentes:
        return None

    feuille = "Ex3 - Mise en forme"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    categories = theme["categories"]
    ventes = [rng.randint(40, 150) for _ in categories]

    for ws in (ws_e, ws_c):
        ws.append(["Categorie", "Ventes"])
        for i, (cat, vente) in enumerate(zip(categories, ventes), start=2):
            ws[f"A{i}"] = cat
            ws[f"B{i}"] = vente
    fin = 1 + len(categories)

    criteres = []
    consignes = []
    if "mise_en_forme_conditionnelle" in actives:
        seuil = round(sum(ventes) / len(ventes))
        plage = f"B2:B{fin}"
        ws_c.conditional_formatting.add(
            plage, CellIsRule(operator="greaterThan", formula=[str(seuil)], fill=None)
        )
        criteres.append({
            "id": "mise_en_forme_conditionnelle", "competence": competences_par_id["mise_en_forme_conditionnelle"]["label"], "description": "Mise en forme conditionnelle appliquee sur les ventes",
            "type": "mise_en_forme_conditionnelle", "plage": plage,
            "points": _points(competences_par_id, "mise_en_forme_conditionnelle"),
        })
        consignes.append(_formater_consigne(competences_par_id["mise_en_forme_conditionnelle"], plage=plage, seuil=seuil))
    if "graphique" in actives:
        opts = (options_par_competence or {}).get("graphique", {})
        type_impose = opts.get("type_impose")      # "histogramme" | "courbes" | "secteurs" | None
        titre_oblige = opts.get("titre_oblige", False)
        tendance = opts.get("tendance", False)

        cls_chart = TYPES_GRAPHIQUE.get(type_impose, BarChart)
        chart = cls_chart()
        chart.add_data(Reference(ws_c, min_col=2, min_row=1, max_row=fin), titles_from_data=True)
        if isinstance(chart, (BarChart, LineChart)):
            chart.set_categories(Reference(ws_c, min_col=1, min_row=2, max_row=fin))
        noms_contexte = {
            "librairie": "Librairie", "pharmacie": "Pharmacie", "sport": "Sport",
            "kinesitherapie": "Kinesitherapie", "centre_readaptation": "Centre de readaptation",
        }
        nom_contexte = noms_contexte.get(contexte, contexte.replace("_", " ").capitalize())
        titre_attendu = f"Ventes par categorie - {nom_contexte}"
        if titre_oblige:
            chart.title = titre_attendu
        if tendance and chart.series and not isinstance(chart, PieChart):
            chart.series[0].trendline = Trendline(trendlineType="linear")
        ws_c.add_chart(chart, "D2")

        # Construction de la consigne : pas de "au choix" si le type est impose
        if type_impose:
            nom_type = NOMS_FR_GRAPHIQUE.get(type_impose, type_impose)
            consigne = f"Creez un graphique de type {nom_type} illustrant les ventes par categorie."
        else:
            consigne = _formater_consigne(competences_par_id["graphique"])
        if titre_oblige:
            consigne += f" Donnez au graphique le titre exactement : \"{titre_attendu}\"."
        if tendance:
            consigne += " Ajoutez une courbe de tendance lineaire sur la serie principale."
        consignes.append(consigne)

        critere_graphique = {
            "id": "graphique", "competence": competences_par_id["graphique"]["label"],
            "description": "Graphique illustrant les ventes par categorie",
            "type": "graphique", "feuille": feuille, "points": _points(competences_par_id, "graphique"),
        }
        if type_impose:
            critere_graphique["type_impose"] = type_impose
        if titre_oblige:
            critere_graphique["titre_oblige"] = True
            critere_graphique["titre_attendu"] = titre_attendu
            critere_graphique["mots_cles_titre"] = ["ventes", contexte.replace("_", " ").lower()]
        if tendance:
            critere_graphique["tendance"] = True
        criteres.append(critere_graphique)

    _ecrire_consignes(ws_e, consignes, cellule="F1")

    return {"id": "ex3", "titre": "Mise en forme et visualisation", "feuille": feuille, "criteres": criteres, "consignes": consignes}


def _construire_groupe_tcd(wb_etudiant, wb_corrige, theme, donnees, competences_par_id, actives):
    if "tcd" not in actives:
        return None

    feuille = "Ex4 - TCD"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    entetes = [theme["colonne_item"], "Categorie", "Prix", "Quantite"]
    for ws in (ws_e, ws_c):
        ws.append(entetes)
        for i, ligne in enumerate(donnees, start=2):
            ws[f"A{i}"] = ligne["nom"]
            ws[f"B{i}"] = ligne["categorie"]
            ws[f"C{i}"] = ligne["prix"]
            ws[f"D{i}"] = ligne["quantite"]

    ws_c["G1"] = (
        "Note enseignant : openpyxl ne sait pas creer de vrai tableau croise dynamique. "
        "Construire le TCD manuellement dans ce corrige avant diffusion."
    )

    consignes = [_formater_consigne(competences_par_id["tcd"])]
    _ecrire_consignes(ws_e, consignes)

    return {
        "id": "ex4", "titre": "Tableau croise dynamique", "feuille": feuille,
        "criteres": [{
            "id": "tcd", "competence": competences_par_id["tcd"]["label"], "description": "Tableau croise dynamique cree (niveau initiation)",
            "type": "tcd", "points": _points(competences_par_id, "tcd"),
        }],
        "consignes": consignes,
    }


def _construire_groupe_synthese(wb_etudiant, wb_corrige, competences_par_id, actives):
    if "synthese_texte" not in actives:
        return None

    feuille = "Ex5 - Synthese"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    question = "Expliquez la methode de tri choisie pour l'exercice 2."
    mots_cles = competences_par_id["synthese_texte"].get("mots_cles", ["croissant", "alphabetique", "ordre"])
    for ws in (ws_e, ws_c):
        ws["A1"] = question
    ws_c["B2"] = "Reponse attendue : tri " + " / ".join(mots_cles) + "."

    consignes = [_formater_consigne(competences_par_id["synthese_texte"], cellule="B2", question=question)]
    _ecrire_consignes(ws_e, consignes, cellule="D1")

    return {
        "id": "ex5", "titre": "Question ouverte", "feuille": feuille,
        "criteres": [{
            "id": "synthese_texte", "competence": competences_par_id["synthese_texte"]["label"], "description": "Reponse texte : explication du choix de tri",
            "type": "texte_motcles", "cellule": "B2", "mots_cles": mots_cles, "mode": "un_de",
            "points": _points(competences_par_id, "synthese_texte"),
        }],
        "consignes": consignes,
    }


def _construire_groupe_recopie(wb_etudiant, wb_corrige, theme, donnees, competences_par_id, rng, actives):
    if "recopie_formule" not in actives:
        return None

    feuille = "Ex6 - Recopie"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    taux = round(rng.uniform(0.05, 0.25), 2)

    for ws in (ws_e, ws_c):
        ws["A1"] = "Taux de remise applique :"
        ws["B1"] = taux
        ws.append([])
        ws.append([theme["colonne_item"], "Prix", "Montant remise"])

    debut = 4
    for i, ligne in enumerate(donnees):
        r = debut + i
        for ws in (ws_e, ws_c):
            ws[f"A{r}"] = ligne["nom"]
            ws[f"B{r}"] = ligne["prix"]
        ws_c[f"C{r}"] = f"=B{r}*$B$1"
    fin = debut + len(donnees) - 1

    cellules = [f"C{debut + i}" for i in range(len(donnees))]
    valeurs_reference = {f"C{debut + i}": round(ligne["prix"] * taux, 2) for i, ligne in enumerate(donnees)}

    comp = competences_par_id["recopie_formule"]
    critere = {
        "id": "recopie_formule", "competence": comp["label"], "description": "Montant de la remise : reference relative + absolue",
        "type": "recopie_formule", "cellules": cellules, "cellule_absolue": "B1", "colonne_relative": "B",
        "valeurs_reference": valeurs_reference, "tolerance": 0.01,
        "points": comp.get("points_defaut", 4), "formule_obligatoire": True,
    }
    consignes = [_formater_consigne(comp, colonne_resultat="C", cellules=f"C{debut}:C{fin}", cellule_absolue="B1")]
    _ecrire_consignes(ws_e, consignes, cellule="E1")

    return {"id": "ex6", "titre": "Recopie de formule", "feuille": feuille, "criteres": [critere], "consignes": consignes}


def _construire_groupe_format_cellule(wb_etudiant, wb_corrige, theme, donnees, competences_par_id, actives):
    if "mise_en_forme_cellule" not in actives:
        return None

    feuille = "Ex7 - Mise en forme cellules"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    entetes = [theme["colonne_item"], "Categorie", "Prix"]
    for ws in (ws_e, ws_c):
        ws.append(entetes)
        for i, ligne in enumerate(donnees, start=2):
            ws[f"A{i}"] = ligne["nom"]
            ws[f"B{i}"] = ligne["categorie"]
            ws[f"C{i}"] = ligne["prix"]

    remplissage = PatternFill(start_color="5BC4E8", end_color="5BC4E8", fill_type="solid")
    for cell in ws_c[1]:
        cell.font = Font(bold=True)
        cell.fill = remplissage

    comp = competences_par_id["mise_en_forme_cellule"]
    plage = "A1:C1"
    critere = {
        "id": "mise_en_forme_cellule", "competence": comp["label"], "description": "Mise en forme de la ligne d'en-tete (gras + couleur)",
        "type": "mise_en_forme_cellule", "plage": plage, "points": comp.get("points_defaut", 2),
    }
    consignes = [_formater_consigne(comp, plage=plage)]
    _ecrire_consignes(ws_e, consignes, cellule="E1")

    return {"id": "ex7", "titre": "Mise en forme d'une feuille de calcul", "feuille": feuille,
            "criteres": [critere], "consignes": consignes}


def _construire_groupe_format_nombre(wb_etudiant, wb_corrige, theme, donnees, competences_par_id, actives):
    if "format_pourcentage" not in actives:
        return None

    feuille = "Ex8 - Formats nombres"
    ws_e, ws_c = wb_etudiant.create_sheet(feuille), wb_corrige.create_sheet(feuille)
    for ws in (ws_e, ws_c):
        ws.append([theme["colonne_item"], "Taux de reussite"])

    debut = 2
    for i, ligne in enumerate(donnees):
        r = debut + i
        taux_reussite = ligne["taux_reussite"]
        for ws in (ws_e, ws_c):
            ws[f"A{r}"] = ligne["nom"]
            ws[f"B{r}"] = taux_reussite
        ws_c[f"B{r}"].number_format = "0%"
    fin = debut + len(donnees) - 1

    cellules = [f"B{debut + i}" for i in range(len(donnees))]
    valeurs_reference = {f"B{debut + i}": ligne["taux_reussite"] for i, ligne in enumerate(donnees)}

    comp = competences_par_id["format_pourcentage"]
    critere = {
        "id": "format_pourcentage", "competence": comp["label"], "description": "Format pourcentage applique sans modifier la valeur",
        "type": "format_nombre", "cellules": cellules, "format_attendu": comp.get("format_attendu", "%"),
        "valeurs_reference": valeurs_reference, "tolerance": 0.001, "points": comp.get("points_defaut", 2),
    }
    consignes = [_formater_consigne(comp, colonne_resultat="B", cellules=f"B{debut}:B{fin}")]
    _ecrire_consignes(ws_e, consignes, cellule="D1")

    return {"id": "ex8", "titre": "Formats de nombres", "feuille": feuille,
            "criteres": [critere], "consignes": consignes}


CELLULE_NOM = "B3"
CELLULE_PRENOM = "B4"


def _construire_page_garde(wb_etudiant, contexte, session, annee, variante, exercices):
    """Insere une feuille de garde : identite de l'etudiant puis consignes de chaque exercice."""
    ws = wb_etudiant.create_sheet("0 - Consignes", 0)
    ws["A1"] = f"Epreuve Excel — session {session} {annee} — variante {variante} — contexte {contexte}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 30

    ws["A3"] = "Nom :"
    ws["A4"] = "Prenom :"
    remplissage = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for cellule in (CELLULE_NOM, CELLULE_PRENOM):
        ws[cellule].fill = remplissage
        ws[cellule].font = Font(bold=True)
    ws["A6"] = "A completer avant de rendre votre copie."
    ws["A6"].font = Font(italic=True, size=9)

    ligne = 8
    for exercice in exercices:
        ws[f"A{ligne}"] = exercice["titre"]
        ws[f"A{ligne}"].font = Font(bold=True, size=12)
        ligne += 1
        for consigne in exercice.get("consignes", []):
            ws[f"A{ligne}"] = f"- {consigne}"
            ws[f"A{ligne}"].alignment = Alignment(wrap_text=True)
            ligne += 1
        ligne += 1


def generer_epreuve_excel(contexte, session, annee, variante, competences_actives, points_par_competence=None,
                           nb_lignes=5, seed=None, competences=None, activer_watermark=False,
                           options_par_competence=None):
    """Genere (wb_etudiant, wb_corrige, config) pour une epreuve Excel.

    `competences_actives` : iterable d'ids de competences a inclure (cf. data/competences.json).
    `points_par_competence` : dict optionnel {id: points} pour surcharger le bareme par defaut du catalogue.
    `activer_watermark` : si True, incruste un identifiant de tracabilite (UUID) dans le
    fichier etudiant et stocke dans la config un hash des zones non editables. Permet de
    detecter une substitution de fichier ou une falsification des donnees source — pas
    de detecter l'usage d'une IA (cf. discussion onglet Generateur).
    """
    competences = competences if competences is not None else charger_competences()
    competences_par_id = {c["id"]: dict(c) for c in competences}
    for comp_id, points in (points_par_competence or {}).items():
        if comp_id in competences_par_id:
            competences_par_id[comp_id]["points_defaut"] = points

    actives = set(competences_actives) & set(competences_par_id)
    if not actives:
        raise ValueError("Aucune competence active selectionnee")

    themes = charger_themes()
    if contexte not in themes:
        raise ValueError(f"Contexte inconnu : {contexte} (disponibles : {list(themes)})")
    theme = themes[contexte]
    rng = random.Random(seed)

    donnees_ex1 = _construire_donnees(theme, nb_lignes, rng)
    donnees_ex2a = _construire_donnees(theme, nb_lignes, rng)
    donnees_ex2b = _construire_donnees(theme, nb_lignes, rng)
    donnees_ex4 = _construire_donnees(theme, nb_lignes, rng)
    donnees_ex6 = _construire_donnees(theme, nb_lignes, rng)
    donnees_ex7 = _construire_donnees(theme, min(nb_lignes, 6), rng)
    donnees_ex8 = _construire_donnees(theme, nb_lignes, rng)

    wb_etudiant = openpyxl.Workbook()
    wb_etudiant.remove(wb_etudiant.active)
    wb_corrige = openpyxl.Workbook()
    wb_corrige.remove(wb_corrige.active)

    exercices = [
        _construire_groupe_formules(wb_etudiant, wb_corrige, theme, donnees_ex1, competences_par_id, actives),
        _construire_groupe_tri(wb_etudiant, wb_corrige, theme, donnees_ex2a, donnees_ex2b, competences_par_id, actives, rng),
        _construire_groupe_visuel(wb_etudiant, wb_corrige, theme, competences_par_id, actives, rng, options_par_competence, contexte),
        _construire_groupe_tcd(wb_etudiant, wb_corrige, theme, donnees_ex4, competences_par_id, actives),
        _construire_groupe_synthese(wb_etudiant, wb_corrige, competences_par_id, actives),
        _construire_groupe_recopie(wb_etudiant, wb_corrige, theme, donnees_ex6, competences_par_id, rng, actives),
        _construire_groupe_format_cellule(wb_etudiant, wb_corrige, theme, donnees_ex7, competences_par_id, actives),
        _construire_groupe_format_nombre(wb_etudiant, wb_corrige, theme, donnees_ex8, competences_par_id, actives),
    ]
    exercices = [e for e in exercices if e is not None]

    _construire_page_garde(wb_etudiant, contexte, session, annee, variante, exercices)

    config = {
        "session": session, "annee": annee, "module": "excel", "variante": variante,
        "contexte": contexte, "exercices": exercices,
    }
    config["bareme_total"] = sum(
        critere.get("points", critere.get("points_par_cellule", 0) * len(critere.get("cellules", [])))
        for exercice in exercices
        for critere in exercice["criteres"]
    )

    if activer_watermark:
        identifiant = generer_uuid()
        definir_watermark_xlsx(wb_etudiant, identifiant)
        config["watermark"] = {
            "id": identifiant,
            "hash_zones_figees": hash_zones_figees(wb_etudiant, config),
        }

    return wb_etudiant, wb_corrige, config
